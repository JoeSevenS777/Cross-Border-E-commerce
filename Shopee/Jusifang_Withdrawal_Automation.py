# -*- coding: utf-8 -*-
"""
Jusifang Withdrawal Automation (Template → Process → Archive → Replenish)

What this script does
---------------------
1) Treat one workbook as the "Master Template" (default layout; normally no embedded images).
2) For each other .xlsx workbook in the same folder:
   - If it has embedded images in column `提款截圖`, run OCR and fill missing fields:
        日期 / 提款編號 / 提款金額
   - Summarize newest date + total amount
   - Save a renamed "finished" workbook:
        yyMMdd + 核心檔名 + total_amount.xlsx
   - Move the finished workbook into a corresponding folder:
        ./核心檔名/finished.xlsx
   - Replenish a fresh default template back to base folder as:
        ./核心檔名.xlsx
     (by copying from the Master Template)

Error handling (your requirement)
---------------------------------
If something goes wrong, the script prints a clear ERROR message, for example:
- Master template missing
- Required headers missing
- Workbook is open / permission denied
- Finished workbook name already exists in destination folder
- The replenished template name already exists and is not safe to overwrite

"""

from __future__ import annotations

import os
import re
import shutil
import datetime as dt
import traceback

import openpyxl
from PIL import Image
import pytesseract


# =========================
# CONFIGURATION
# =========================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Template sources folder (per-core). After processing a workbook, we copy a fresh template
# from ./_template_sources/<核心檔名>.xlsx back into BASE_DIR as <核心檔名>.xlsx
#
# The script will AUTO-LEARN these sources:
# - If it sees a workbook in BASE_DIR with required headers AND with NO images in 提款截圖,
#   it will copy it into _template_sources as the per-core template source.
TEMPLATE_SOURCES_DIRNAME = "_template_sources"

# Prefer this sheet name; fallback to first sheet if not present
SHEET_NAME = "Sheet1"

# Tesseract (adjust if needed)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESS_LANG = "chi_tra+eng"   # if chi_tra not installed, try "eng"


# =========================
# UTILITIES
# =========================

def log_info(msg: str) -> None:
    print(msg)


def log_error(msg: str) -> None:
    print(f"ERROR: {msg}")


def safe_makedirs(path: str) -> bool:
    try:
        os.makedirs(path, exist_ok=True)
        return True
    except Exception as e:
        log_error(f"Cannot create folder: {path} ({e})")
        return False


# =========================
# OCR + PARSING HELPERS
# =========================

def ocr_image(pil_image: Image.Image) -> str:
    return pytesseract.image_to_string(pil_image, lang=TESS_LANG)


def parse_date_from_text(text: str) -> dt.date | None:
    m = re.search(r"(20\d{2}/\d{2}/\d{2})", text)
    if not m:
        return None
    date_str = m.group(1)
    try:
        return dt.datetime.strptime(date_str, "%Y/%m/%d").date()
    except ValueError:
        return None


def parse_withdraw_id_from_text(text: str) -> str | None:
    """Extract 提款編號, allowing spaces inserted by OCR."""

    def looks_like_date8(s: str) -> bool:
        if len(s) != 8 or not s.isdigit():
            return False
        year = int(s[:4]); month = int(s[4:6]); day = int(s[6:8])
        return 2000 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31

    # Prefer labeled capture
    m = re.search(r"提款編號\s*[:：]?\s*([0-9][0-9\s]{14,30})", text)
    if m:
        raw = m.group(1)
        digits = re.sub(r"\s+", "", raw)
        if 16 <= len(digits) <= 22 and digits.isdigit() and not looks_like_date8(digits):
            return digits

    # Fallback: any 16–22 digit chunk (allow spaces)
    candidates: list[str] = []
    for mm in re.finditer(r"(?<!\d)(?:\d\s*){16,22}(?!\d)", text):
        digits = re.sub(r"\s+", "", mm.group(0))
        if digits.isdigit() and not looks_like_date8(digits):
            candidates.append(digits)

    if candidates:
        candidates.sort(key=len, reverse=True)
        return candidates[0]
    return None


def parse_amount_from_text(text: str) -> int | None:
    """Extract withdrawal amount. Target is negative on slips, return positive int."""

    KEYWORDS = ("提領總額", "提款金額", "提領金額")
    lines = text.splitlines()

    # Keyword line with explicit negative amount
    for line in lines:
        clean = line.replace(" ", "")
        if any(k in clean for k in KEYWORDS):
            m = re.search(r"[-−]\s*(?:NT\$|NT|新台幣)?\s*([0-9][0-9,\.]*)", line)
            if m:
                num_str = m.group(1).replace(",", "").strip()
                if "." in num_str:
                    num_str = num_str.split(".", 1)[0]
                try:
                    val = int(num_str)
                except ValueError:
                    continue
                if val != 0:
                    return abs(val)

    def looks_like_date8(s: str) -> bool:
        if len(s) != 8 or not s.isdigit():
            return False
        year = int(s[:4]); month = int(s[4:6]); day = int(s[6:8])
        return 2000 <= year <= 2099 and 1 <= month <= 12 and 1 <= day <= 31

    # Fallback: scan all negative amounts
    candidates: list[int] = []
    for m in re.finditer(r"[-−]\s*(?:NT\$|NT|新台幣)?\s*([0-9][0-9,\.]*)", text):
        num_str = m.group(1).replace(",", "").strip()
        if not num_str:
            continue
        if len(num_str) > 9:      # likely an ID
            continue
        if looks_like_date8(num_str):
            continue
        if "." in num_str:
            num_str = num_str.split(".", 1)[0]
        try:
            val = int(num_str)
        except ValueError:
            continue
        if val == 0 or val > 2_000_000:
            continue
        candidates.append(val)

    if candidates:
        best = max(candidates, key=lambda v: abs(v))
        return abs(best)
    return None


def extract_fields_from_image(pil_image: Image.Image) -> tuple[dt.date | None, str | None, int | None]:
    text = ocr_image(pil_image)
    return (
        parse_date_from_text(text),
        parse_withdraw_id_from_text(text),
        parse_amount_from_text(text),
    )


# =========================
# SHEET HELPERS
# =========================

REQUIRED_HEADERS = ("日期", "提款編號", "提款金額", "提款截圖")


def get_target_worksheet(wb: openpyxl.Workbook):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    return wb[wb.sheetnames[0]]


def get_header_map(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def normalize_date_cell(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
            try:
                return dt.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def normalize_amount_cell(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.replace(",", "").strip()
        if s == "":
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def summarize_sheet(ws, col_date: int, col_amount: int) -> tuple[dt.date | None, int]:
    dates: list[dt.date] = []
    total = 0
    for r in range(2, ws.max_row + 1):
        d = normalize_date_cell(ws.cell(row=r, column=col_date).value)
        if d:
            dates.append(d)
        amt = normalize_amount_cell(ws.cell(row=r, column=col_amount).value)
        if amt is not None:
            total += amt
    return (max(dates) if dates else None), total


def core_name_from_filename(stem: str) -> str:
    if re.match(r"^\d{6}", stem):
        stem = stem[6:]
    stem = re.sub(r"\d+$", "", stem)
    stem = stem.strip(" -_")
    return stem or "個人卡提"


def workbook_has_images_in_column(ws, col_image: int) -> bool:
    imgs = getattr(ws, "_images", [])
    for img in imgs:
        try:
            anchor_from = img.anchor._from  # 0-based row/col
            col_index = anchor_from.col + 1
        except Exception:
            continue
        if col_index == col_image:
            return True
    return False


# =========================
# TEMPLATE SOURCE + REPLENISH
# =========================

def get_template_sources_dir() -> str:
    return os.path.join(BASE_DIR, TEMPLATE_SOURCES_DIRNAME)


def ensure_template_sources_dir() -> bool:
    return safe_makedirs(get_template_sources_dir())


def template_source_path_for_core(core: str) -> str:
    return os.path.join(get_template_sources_dir(), f"{core}.xlsx")


def learn_template_source_if_applicable(xlsx_path: str) -> None:
    """If a workbook in BASE_DIR looks like a clean default template (no images in 提款截圖),
    store a per-core template source copy in _template_sources.

    This fixes the issue where 馬/白 have different default data in other columns: we keep
    one template source per core name.
    """
    filename = os.path.basename(xlsx_path)
    stem, _ = os.path.splitext(filename)
    core = core_name_from_filename(stem)

    # skip anything already inside the sources dir
    if os.path.abspath(os.path.dirname(xlsx_path)) == os.path.abspath(get_template_sources_dir()):
        return

    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = get_target_worksheet(wb)
        headers = get_header_map(ws)
        if any(h not in headers for h in REQUIRED_HEADERS):
            return
        if workbook_has_images_in_column(ws, headers["提款截圖"]):
            return
    except Exception:
        return

    src_path = template_source_path_for_core(core)
    if os.path.exists(src_path):
        return

    try:
        shutil.copy2(xlsx_path, src_path)
        log_info(f"Learned template source: {TEMPLATE_SOURCES_DIRNAME}/{core}.xlsx (from {filename})")
    except Exception as e:
        log_error(f"Failed to learn template source from '{filename}': {e}")


def replenish_template_from_source(core: str) -> bool:
    """Revert BASE_DIR/<core>.xlsx back to its default version from the per-core template source."""
    src = template_source_path_for_core(core)
    dst = os.path.join(BASE_DIR, f"{core}.xlsx")

    if not os.path.exists(src):
        log_error(f"Template source missing for core '{core}': {TEMPLATE_SOURCES_DIRNAME}/{core}.xlsx")
        log_error("Fix: Put the correct default template workbook in the base folder once (with NO images), rerun, and the script will auto-save it into _template_sources.")
        return False

    # Safety: if dst exists and contains images, refuse overwrite
    if os.path.exists(dst):
        try:
            wb = openpyxl.load_workbook(dst)
            ws = get_target_worksheet(wb)
            headers = get_header_map(ws)
            col_img = headers.get("提款截圖")
            if col_img and workbook_has_images_in_column(ws, col_img):
                log_error(f"Cannot revert because '{core}.xlsx' exists and contains images (working file).")
                return False
        except PermissionError:
            log_error(f"Cannot open existing '{core}.xlsx' (it may be open). Close it and rerun.")
            return False
        except Exception:
            log_error(f"Cannot safely verify existing '{core}.xlsx'. Not overwriting.")
            return False

        # overwrite is OK (we want a real revert), but only if not open
        try:
            os.remove(dst)
        except PermissionError:
            log_error(f"Cannot overwrite '{core}.xlsx' (it may be open). Close it and rerun.")
            return False
        except Exception as e:
            log_error(f"Cannot remove existing '{core}.xlsx' to revert: {e}")
            return False

    try:
        shutil.copy2(src, dst)
        log_info(f"Reverted template: {core}.xlsx (from {TEMPLATE_SOURCES_DIRNAME}/{core}.xlsx)")
        return True
    except PermissionError:
        log_error(f"Cannot write '{core}.xlsx' (it may be open). Close it and rerun.")
        return False
    except Exception as e:
        log_error(f"Failed to revert template '{core}.xlsx': {e}")
        return False


# =========================
# PROCESS ONE WORKBOOK
# =========================

def process_workbook(filepath: str, _unused: object = None) -> None:
    filename = os.path.basename(filepath)
    stem, ext = os.path.splitext(filename)

    log_info(f"\n=== Processing: {filename} ===")

    # Load workbook
    try:
        wb = openpyxl.load_workbook(filepath)
    except PermissionError:
        log_error(f"Cannot open '{filename}' (likely open in Excel). Close it and rerun.")
        return
    except Exception as e:
        log_error(f"Failed to open '{filename}': {e}")
        return

    ws = get_target_worksheet(wb)
    headers = get_header_map(ws)

    # Validate headers
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        log_error(f"Missing required headers in '{filename}': {', '.join(missing)}")
        return

    col_date = headers["日期"]
    col_id = headers["提款編號"]
    col_amount = headers["提款金額"]
    col_image = headers["提款截圖"]

    # If no images, treat as template/empty, skip (no error)
    if not workbook_has_images_in_column(ws, col_image):
        log_info("No embedded images found in 提款截圖 → skipped (template/empty workbook).")
        return

    # OCR fill
    filled_rows = 0
    failed_rows = 0

    for img in getattr(ws, "_images", []):
        try:
            anchor_from = img.anchor._from
            row_index = anchor_from.row + 1
            col_index = anchor_from.col + 1
        except Exception:
            continue

        if col_index != col_image or row_index == 1:
            continue

        v_date = ws.cell(row=row_index, column=col_date).value
        v_id = ws.cell(row=row_index, column=col_id).value
        v_amt = ws.cell(row=row_index, column=col_amount).value

        if v_date not in (None, "") and v_id not in (None, "") and v_amt not in (None, ""):
            continue

        try:
            pil_img = Image.open(img.ref)
            date_value, withdraw_id, amount = extract_fields_from_image(pil_img)

            if date_value and v_date in (None, ""):
                ws.cell(row=row_index, column=col_date).value = date_value
            if withdraw_id and v_id in (None, ""):
                ws.cell(row=row_index, column=col_id).value = str(withdraw_id)
            if amount is not None and v_amt in (None, ""):
                ws.cell(row=row_index, column=col_amount).value = amount

            filled_rows += 1
            log_info(f"  Row {row_index}: date={date_value}, id={withdraw_id}, amount={amount}")

        except Exception as e:
            failed_rows += 1
            log_error(f"  Row {row_index}: OCR/parse failed ({e})")

    newest_date, total_amount = summarize_sheet(ws, col_date, col_amount)

    if not newest_date or total_amount <= 0:
        log_error("Cannot rename/archive because newest_date or total_amount is missing/invalid.")
        log_error("Please confirm the screenshots contain readable date and negative withdrawal amount.")
        return

    date_prefix = newest_date.strftime("%y%m%d")
    core = core_name_from_filename(stem)
    new_stem = f"{date_prefix}{core}{total_amount}"
    finished_filename = new_stem + ext

    dest_folder = os.path.join(BASE_DIR, core)
    if not safe_makedirs(dest_folder):
        return

    dest_path = os.path.join(dest_folder, finished_filename)

    if os.path.exists(dest_path):
        log_error(f"Finished workbook already exists: {core}/{finished_filename}")
        log_error("To avoid overwriting, the script will stop for this file.")
        return

    # Save to a temporary path in base dir first (reduces cross-volume issues)
    tmp_path = os.path.join(BASE_DIR, finished_filename)

    try:
        wb.save(tmp_path)
    except PermissionError:
        log_error(f"Cannot save '{finished_filename}' (likely open). Close it and rerun.")
        return
    except Exception as e:
        log_error(f"Failed to save '{finished_filename}': {e}")
        return

    # Move to destination folder
    try:
        shutil.move(tmp_path, dest_path)
        log_info(f"Archived finished workbook → {core}/{finished_filename}")
    except PermissionError:
        log_error(f"Cannot move '{finished_filename}' to folder '{core}' (file/folder may be open).")
        # Try to keep tmp_path in place; do not delete original
        return
    except Exception as e:
        log_error(f"Failed to move finished workbook: {e}")
        return

    # Delete the original working workbook (the one with embedded images)
    try:
        os.remove(filepath)
        log_info(f"Removed working file: {filename}")
    except PermissionError:
        log_error(f"Cannot delete working file '{filename}' (likely open). Please delete it manually.")
    except Exception as e:
        log_error(f"Could not delete working file '{filename}': {e}")

    # Replenish a fresh template in base folder: core.xlsx
    replenish_template_from_source(core)

    log_info(f"Done. Filled rows={filled_rows}, failed rows={failed_rows}, total={total_amount}, newest_date={newest_date}")


# =========================
# ENTRY POINT
# =========================

def main() -> None:
    # Ensure template sources folder exists
    if not ensure_template_sources_dir():
        return

    xlsx_files = [
        f for f in os.listdir(BASE_DIR)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    ]

    if not xlsx_files:
        log_info("No .xlsx files found in folder.")
        return

    # 1) Learn template sources from any clean (no-image) workbooks in BASE_DIR.
    #    This prevents '馬' reverting to '白'. Each core gets its own template source.
    for fname in xlsx_files:
        learn_template_source_if_applicable(os.path.join(BASE_DIR, fname))

    # 2) Process all workbooks (templates will be auto-skipped because they have no images).
    log_info(f"Found {len(xlsx_files)} .xlsx file(s) in: {BASE_DIR}")

    for fname in xlsx_files:
        try:
            process_workbook(os.path.join(BASE_DIR, fname))
        except Exception as e:
            log_error(f"Unexpected failure processing '{fname}': {e}")
            log_error(traceback.format_exc())


if __name__ == "__main__":
    main()

