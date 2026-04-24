from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

try:
    import fitz  # PyMuPDF
except ImportError:
    print("Missing package: PyMuPDF. Install it with: pip install pymupdf")
    raise

try:
    import openpyxl
except ImportError:
    print("Missing package: openpyxl. Install it with: pip install openpyxl")
    raise

# Put this script and the PDF in:
# D:\JoeProgramFiles\Replace Picklist Shelf in Logistics Label
SCRIPT_FOLDER = Path(__file__).resolve().parent

PDF_INPUT = None
WORKBOOK_INPUT = r"D:\JoeProgramFiles\BigSeller库存管理\台湾货架位明细表_含打印预览（现版本）.xlsm"
PDF_OUTPUT = None

WORKBOOK_SHEET_NAME = None
WORKBOOK_SKU_COLUMN = "SKU编号"
WORKBOOK_SHELF_COLUMN = "台湾货架位"

FONT_NAME = "helv"
NEW_SHELF_FONT_SIZE = 6.8
SHELF_PATTERN = re.compile(r"^[A-Z]+-\d+$", re.I)


@dataclass
class PickRow:
    old_shelf: str
    sku: str
    qty: str
    y0: float
    y1: float
    shelf_rect: fitz.Rect
    row_words: list[dict]


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    replacements = [
        ("\u000b", ""),
        ("\ufffe", ""),
        ("\ufeff", ""),
        ("\xa0", ""),
        ("　", ""),
        (" ", ""),
        ("\t", ""),
        ("\r", ""),
        ("\n", ""),
        ("；", ";"),
        ("，", ","),
        ("：", ":"),
        ("－", "-"),
        ("—", "-"),
        ("–", "-"),
    ]
    for old, new in replacements:
        text = text.replace(old, new)
    return text.strip()


def find_default_pdf(script_folder: Path) -> Path:
    pdf_files = sorted(
        [p for p in script_folder.glob("*.pdf") if not p.name.startswith("output_")],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not pdf_files:
        raise FileNotFoundError(
            f"No PDF file found in script folder: {script_folder}\n"
            "Put the logistics label PDF in the same folder as this script."
        )
    if len(pdf_files) > 1:
        print("Multiple PDF files found. The newest one will be used:")
        for p in pdf_files:
            print(f"  {p.name}")
    return pdf_files[0]


def build_default_output_path(pdf_path: Path) -> Path:
    return pdf_path.with_name(f"output_台湾货架位_{pdf_path.stem}.pdf")


def choose_workbook_sheet(wb):
    if WORKBOOK_SHEET_NAME:
        return wb[WORKBOOK_SHEET_NAME]

    target = normalize_text(WORKBOOK_SHELF_COLUMN)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
            values = [normalize_text(cell.value) for cell in row]
            if target in values:
                return ws
    raise ValueError(f"No sheet contains column: {WORKBOOK_SHELF_COLUMN}")


def find_header_row_and_columns(ws):
    sku_header = normalize_text(WORKBOOK_SKU_COLUMN)
    shelf_header = normalize_text(WORKBOOK_SHELF_COLUMN)

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        values = [normalize_text(cell.value) for cell in row]
        if sku_header in values and shelf_header in values:
            return row[0].row, values.index(sku_header) + 1, values.index(shelf_header) + 1, values

    raise ValueError(
        f"Cannot find required columns. Need: {WORKBOOK_SKU_COLUMN}, {WORKBOOK_SHELF_COLUMN}"
    )


def load_sku_to_shelf_map(workbook_path: str) -> dict[str, str]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    ws = choose_workbook_sheet(wb)
    header_row, sku_col, shelf_col, headers = find_header_row_and_columns(ws)

    lookup: dict[str, str] = {}
    duplicate_count = 0

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        sku = normalize_text(row[sku_col - 1].value)
        shelf = normalize_text(row[shelf_col - 1].value)
        if not sku or not shelf:
            continue
        if sku in lookup and lookup[sku] != shelf:
            duplicate_count += 1
            continue
        lookup[sku] = shelf

    print(f"Workbook sheet: {ws.title}")
    print(f"Workbook SKU column: {headers[sku_col - 1]}")
    print(f"Workbook shelf column: {headers[shelf_col - 1]}")
    print(f"Loaded SKU → 台湾货架位 records: {len(lookup)}")
    if duplicate_count:
        print(f"WARNING: duplicate SKU rows ignored: {duplicate_count}")
    return lookup


def page_is_pick_list(page) -> bool:
    text = page.get_text("text")
    return all(k in text for k in ["Shelf", "SKU", "QTY", "GTIN"])


def get_words(page) -> list[dict]:
    result = []
    for w in page.get_text("words"):
        x0, y0, x1, y1, text, block_no, line_no, word_no = w
        result.append({
            "x0": x0,
            "y0": y0,
            "x1": x1,
            "y1": y1,
            "text": text,
            "rect": fitz.Rect(x0, y0, x1, y1),
        })
    return result


def find_header_positions(words: list[dict]):
    headers = [w for w in words if w["text"] in {"Shelf", "SKU", "QTY", "GTIN"}]
    if len(headers) < 4:
        raise ValueError("Cannot locate pick-list headers.")

    shelf_h = min([w for w in headers if w["text"] == "Shelf"], key=lambda w: w["x0"])
    sku_h = min([w for w in headers if w["text"] == "SKU"], key=lambda w: w["x0"])
    qty_h = min([w for w in headers if w["text"] == "QTY"], key=lambda w: w["x0"])
    gtin_h = min([w for w in headers if w["text"] == "GTIN"], key=lambda w: w["x0"])

    header_bottom = max(w["y1"] for w in headers)
    shelf_sku_boundary = (shelf_h["x1"] + sku_h["x0"]) / 2
    sku_qty_boundary = (sku_h["x1"] + qty_h["x0"]) / 2
    qty_gtin_boundary = (qty_h["x1"] + gtin_h["x0"]) / 2

    return header_bottom, shelf_sku_boundary, sku_qty_boundary, qty_gtin_boundary


def extract_pick_rows(page) -> list[PickRow]:
    words = get_words(page)
    header_bottom, shelf_sku_boundary, sku_qty_boundary, qty_gtin_boundary = find_header_positions(words)

    body_words = [w for w in words if w["y0"] > header_bottom + 2]

    shelf_words = sorted(
        [w for w in body_words if SHELF_PATTERN.match(w["text"].strip()) and w["x1"] < shelf_sku_boundary + 20],
        key=lambda w: w["y0"],
    )

    rows: list[PickRow] = []
    for i, shelf_word in enumerate(shelf_words):
        y0 = max(header_bottom, shelf_word["y0"] - 8)
        y1 = shelf_words[i + 1]["y0"] - 8 if i + 1 < len(shelf_words) else min(page.rect.height, shelf_word["y1"] + 60)
        if y1 <= y0:
            y1 = shelf_word["y1"] + 30

        row_words = [w for w in body_words if y0 <= (w["y0"] + w["y1"]) / 2 < y1]

        sku_words = sorted(
            [w for w in row_words if shelf_sku_boundary <= w["x0"] < sku_qty_boundary],
            key=lambda w: (w["y0"], w["x0"]),
        )
        qty_words = sorted(
            [w for w in row_words if sku_qty_boundary <= w["x0"] < qty_gtin_boundary],
            key=lambda w: (w["y0"], w["x0"]),
        )

        sku = normalize_text("".join(w["text"] for w in sku_words))
        qty = normalize_text("".join(w["text"] for w in qty_words))

        rows.append(PickRow(
            old_shelf=normalize_text(shelf_word["text"]),
            sku=sku,
            qty=qty,
            y0=y0,
            y1=y1,
            shelf_rect=shelf_word["rect"],
            row_words=row_words,
        ))

    return rows


def find_shelf_for_sku(sku: str, lookup: dict[str, str]) -> tuple[str | None, str]:
    key = normalize_text(sku)
    if key in lookup:
        return lookup[key], "exact"

    contained = [(k, v) for k, v in lookup.items() if key and key in k]
    if len(contained) == 1:
        return contained[0][1], "pdf_sku_contained_in_workbook_sku"

    reverse = [(k, v) for k, v in lookup.items() if k and k in key]
    if len(reverse) == 1:
        return reverse[0][1], "workbook_sku_contained_in_pdf_sku"

    if len(contained) > 1 or len(reverse) > 1:
        return None, "ambiguous"
    return None, "not_found"


def draw_white_box(page, rect: fitz.Rect):
    page.draw_rect(rect, color=(1, 1, 1), fill=(1, 1, 1), width=0, overlay=True)


def replace_shelf_at_original_position(page, row: PickRow, new_shelf: str):
    old = row.shelf_rect

    # Cover only the old shelf text area. Keep the table grid and other columns unchanged.
    cover = fitz.Rect(old.x0 - 2, old.y0 - 2, old.x0 + 64, old.y1 + 4)
    draw_white_box(page, cover)

    # No wrapping: write the full shelf code on one line.
    page.insert_text(
        (old.x0, old.y1),
        new_shelf,
        fontsize=NEW_SHELF_FONT_SIZE,
        fontname=FONT_NAME,
        color=(0, 0, 0),
        overlay=True,
    )


def blank_row_words(page, row: PickRow):
    for w in row.row_words:
        r = w["rect"]
        cover = fitz.Rect(r.x0 - 1, r.y0 - 1, r.x1 + 1, r.y1 + 1)
        draw_white_box(page, cover)


def process_pdf(pdf_path: str, output_path: str, sku_to_shelf: dict[str, str]):
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    pick_pages = 0
    replaced_count = 0
    blanked_count = 0

    for page_index, page in enumerate(doc, start=1):
        if not page_is_pick_list(page):
            continue

        pick_pages += 1
        rows = extract_pick_rows(page)

        for row in rows:
            new_shelf, status = find_shelf_for_sku(row.sku, sku_to_shelf)
            if new_shelf:
                replace_shelf_at_original_position(page, row, new_shelf)
                replaced_count += 1
                print(f"Page {page_index}: REPLACE | SKU={row.sku} | {row.old_shelf} -> {new_shelf} | {status}")
            else:
                blank_row_words(page, row)
                blanked_count += 1
                print(f"Page {page_index}: DELETE/BLANK | SKU={row.sku} | old shelf={row.old_shelf} | reason={status}")

    doc.save(output_path, garbage=4, deflate=True)
    doc.close()

    print("-" * 80)
    print(f"Input PDF: {pdf_path}")
    print(f"Output PDF: {output_path}")
    print(f"Total pages: {total_pages}")
    print(f"Pick-list pages processed: {pick_pages}")
    print(f"Rows replaced: {replaced_count}")
    print(f"Rows blanked because SKU was not found or ambiguous: {blanked_count}")


def main():
    pdf_path = find_default_pdf(SCRIPT_FOLDER) if PDF_INPUT is None else Path(PDF_INPUT)
    workbook_path = Path(WORKBOOK_INPUT)
    output_path = build_default_output_path(pdf_path) if PDF_OUTPUT is None else Path(PDF_OUTPUT)

    if len(sys.argv) >= 2:
        pdf_path = Path(sys.argv[1])
        output_path = build_default_output_path(pdf_path) if PDF_OUTPUT is None else Path(PDF_OUTPUT)
    if len(sys.argv) >= 3:
        workbook_path = Path(sys.argv[2])
    if len(sys.argv) >= 4:
        output_path = Path(sys.argv[3])

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    print(f"Script folder: {SCRIPT_FOLDER}")
    print(f"PDF input: {pdf_path}")
    print(f"Workbook input: {workbook_path}")
    print(f"PDF output: {output_path}")
    print("-" * 80)

    sku_to_shelf = load_sku_to_shelf_map(str(workbook_path))
    process_pdf(str(pdf_path), str(output_path), sku_to_shelf)


if __name__ == "__main__":
    main()
