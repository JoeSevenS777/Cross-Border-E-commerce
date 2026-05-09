"""
Migrate embedded screenshot records from `pics` to `reconciliation`.

Updated behavior:
1. Reads embedded images from the `pics` sheet.
2. Migrates only these direct fields:
   - A: account
   - B: start date
   - C: end date
   - D: revenue
3. Copies formulas/styles from the previous rows like Excel drag-down.
4. Merges J:O by batch, where one batch means same start date + end date.
5. Fills/copies J:N from the previous batch.
6. Leaves O merged but empty, because confirmation is manual.
7. If the new start date does not equal previous end date + 1 day, fills B:C red.
8. Sorts detected image batches by date order before writing.
9. Deletes processed images from `pics` after successful migration.
10. Saves the updated workbook with timestamp suffix: yyyymmddhhmmss.
11. Moves the previous workbook into a `log` folder after successful save.
12. If no workbook argument is supplied, automatically finds the newest matching workbook in the current folder.

Requirements:
    pip install openpyxl pillow pytesseract

Typical usage:
    python migrate_sheet_pics_to_reconciliation.py

Or specify a workbook manually:
    python migrate_sheet_pics_to_reconciliation.py "Shop1&4reconciliation_20260509153022.xlsx"
"""

from __future__ import annotations

import io
import re
import shutil
import sys
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from PIL import Image
import pytesseract


# =========================
# User settings
# =========================

# The script will try these sheet names automatically.
SHEET_PICS_ALIASES = [
    "pics",
    "Pics",
    "Sheet Pics",
    "SheetPics",
    "Sheet_Pics",
    "圖片",
]

SHEET_RECONCILIATION_ALIASES = [
    "reconciliation",
    "Reconciliation",
    "Sheet Reconciliation",
    "Recon",
    "recon",
]

# The script will auto-detect the newest workbook matching this pattern if no workbook is passed.
# It ignores temporary Excel files beginning with "~$" and ignores files inside the log folder.
WORKBOOK_NAME_PATTERN = "Shop1&4reconciliation*.xlsx"
LOG_FOLDER_NAME = "log"

# Current known account mapping.
# Add more accounts here later if needed.
ACCOUNT_MAP: Dict[str, str] = {
    "華南商業銀行 (7494)": "ck1808052",
    "華南商業銀行": "ck1808052",
    "7494": "ck1808052",
    "中華郵政股份有限公司 (6260)": "noso1981",
    "中華郵政股份有限公司": "noso1981",
    "6260": "noso1981",
}

# Preferred output order inside each batch.
ACCOUNT_ORDER = ["ck1808052", "noso1981"]

# Reconciliation columns.
COL_ACCOUNT = 1       # A
COL_START_DATE = 2    # B
COL_END_DATE = 3      # C
COL_REVENUE = 4       # D
MERGE_START_COL = 10  # J
MERGE_END_COL = 15    # O
CONFIRM_COL = 15      # O

HEADER_ROW = 1
FIRST_DATA_ROW = 2

DATE_WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")

# Try Chinese Traditional + English first. If chi_tra is not installed, the script falls back to English.
OCR_LANG = "chi_tra+eng"


@dataclass(frozen=True)
class ImageResult:
    image_index: int
    account: str
    start_date: date
    end_date: date
    revenue: int
    dates: Tuple[date, ...]
    amounts: Tuple[int, ...]
    raw_text: str


# =========================
# OCR and parsing
# =========================

def normalize_text(text: str) -> str:
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def image_to_text(image: Image.Image) -> str:
    """Run OCR on one image."""
    w, h = image.size
    if w < 1800:
        image = image.resize((w * 2, h * 2))

    try:
        return pytesseract.image_to_string(image, lang=OCR_LANG)
    except pytesseract.TesseractError:
        return pytesseract.image_to_string(image, lang="eng")


def identify_account(text: str) -> Optional[str]:
    normalized = normalize_text(text)
    for bank_text, account in ACCOUNT_MAP.items():
        if bank_text in normalized:
            return account
    return None


def extract_dates(text: str) -> List[date]:
    pattern = r"\b(20\d{2})[-/](\d{1,2})[-/](\d{1,2})\b"
    results: List[date] = []
    for y, m, d in re.findall(pattern, text):
        try:
            results.append(date(int(y), int(m), int(d)))
        except ValueError:
            pass
    return results


def extract_amounts(text: str) -> List[int]:
    """Extract negative outgoing amounts and return positive integers."""
    normalized = text.replace("，", ",")
    pattern = r"-\s*(\d{1,3}(?:,\d{3})+|\d{4,})\b"
    amounts: List[int] = []
    for value in re.findall(pattern, normalized):
        try:
            amounts.append(abs(int(value.replace(",", ""))))
        except ValueError:
            pass
    return amounts


def parse_image_result(image_index: int, image: Image.Image) -> ImageResult:
    raw_text = image_to_text(image)
    normalized = normalize_text(raw_text)

    account = identify_account(normalized)
    if not account:
        raise ValueError(
            f"Image {image_index}: Cannot identify account. "
            f"Please add the bank text to ACCOUNT_MAP.\nOCR text:\n{raw_text}"
        )

    dates = extract_dates(normalized)
    if not dates:
        raise ValueError(f"Image {image_index}: Cannot extract dates.\nOCR text:\n{raw_text}")

    amounts = extract_amounts(normalized)
    if not amounts:
        raise ValueError(f"Image {image_index}: Cannot extract negative amounts.\nOCR text:\n{raw_text}")

    return ImageResult(
        image_index=image_index,
        account=account,
        start_date=min(dates),
        end_date=max(dates),
        revenue=sum(amounts),
        dates=tuple(sorted(set(dates))),
        amounts=tuple(amounts),
        raw_text=raw_text,
    )


# =========================
# Workbook helpers
# =========================

def get_sheet_by_alias(wb, aliases: List[str]):
    """Find sheet by exact, case-insensitive, or normalized alias."""
    for alias in aliases:
        if alias in wb.sheetnames:
            return wb[alias]

    lowered = {name.lower(): name for name in wb.sheetnames}
    for alias in aliases:
        matched = lowered.get(alias.lower())
        if matched:
            return wb[matched]

    def norm(s: str) -> str:
        return re.sub(r"[\s_\-]+", "", s).lower()

    normalized = {norm(name): name for name in wb.sheetnames}
    for alias in aliases:
        matched = normalized.get(norm(alias))
        if matched:
            return wb[matched]

    raise ValueError(
        "Cannot find required sheet.\n"
        f"Available sheets: {', '.join(wb.sheetnames)}\n"
        f"Tried aliases: {', '.join(aliases)}"
    )


def get_last_data_row(ws) -> int:
    for row in range(ws.max_row, FIRST_DATA_ROW - 1, -1):
        if ws.cell(row=row, column=COL_ACCOUNT).value not in (None, ""):
            return row
    return HEADER_ROW


def cell_date_value(cell: Cell) -> Optional[date]:
    value = cell.value
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", value.strip())
        if match:
            y, m, d = match.groups()
            return date(int(y), int(m), int(d))
    return None


def existing_row_map(ws) -> Dict[Tuple[str, date, date], int]:
    result: Dict[Tuple[str, date, date], int] = {}
    last_row = get_last_data_row(ws)
    for row in range(FIRST_DATA_ROW, last_row + 1):
        account = ws.cell(row=row, column=COL_ACCOUNT).value
        start = cell_date_value(ws.cell(row=row, column=COL_START_DATE))
        end = cell_date_value(ws.cell(row=row, column=COL_END_DATE))
        if account and start and end:
            result[(str(account), start, end)] = row
    return result


def copy_cell_with_formula_translation(src: Cell, dst: Cell) -> None:
    """Copy value/formula/style from source cell to destination cell, translating formulas."""
    if src.data_type == "f" and isinstance(src.value, str):
        try:
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        except Exception:
            dst.value = src.value
    else:
        dst.value = src.value

    if src.has_style:
        dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)


def copy_row_like_drag_down(ws, template_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height
    for col in range(1, ws.max_column + 1):
        copy_cell_with_formula_translation(ws.cell(template_row, col), ws.cell(target_row, col))


def clear_row_transfer_area(ws, row: int) -> None:
    """Clear J:O values before recreating batch-level merge/formulas."""
    for col in range(MERGE_START_COL, MERGE_END_COL + 1):
        ws.cell(row=row, column=col).value = None


def unmerge_cells_intersecting_rows(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        if (
            merged_range.min_row <= end_row
            and merged_range.max_row >= start_row
            and merged_range.min_col <= end_col
            and merged_range.max_col >= start_col
        ):
            ranges_to_unmerge.append(str(merged_range))

    for range_string in ranges_to_unmerge:
        ws.unmerge_cells(range_string)


def find_previous_batch_top_row(ws, before_row: int) -> Optional[int]:
    """
    Find the top row of the previous batch before `before_row`.
    A batch is identified by same start date + end date.
    """
    if before_row <= FIRST_DATA_ROW:
        return None

    previous_row = before_row - 1
    previous_start = cell_date_value(ws.cell(previous_row, COL_START_DATE))
    previous_end = cell_date_value(ws.cell(previous_row, COL_END_DATE))
    if not previous_start or not previous_end:
        return previous_row

    top = previous_row
    for row in range(previous_row - 1, FIRST_DATA_ROW - 1, -1):
        row_start = cell_date_value(ws.cell(row, COL_START_DATE))
        row_end = cell_date_value(ws.cell(row, COL_END_DATE))
        if row_start == previous_start and row_end == previous_end:
            top = row
        else:
            break
    return top


def setup_transfer_area_for_batch(ws, start_row: int, end_row: int, template_batch_top_row: Optional[int]) -> None:
    """
    Merge J:O vertically by batch.

    J:N are copied from the previous batch top row.
    O is merged but left empty for manual confirmation.
    """
    if end_row < start_row:
        return

    unmerge_cells_intersecting_rows(ws, start_row, end_row, MERGE_START_COL, MERGE_END_COL)

    for col in range(MERGE_START_COL, MERGE_END_COL + 1):
        # Copy style/formula/value to the top cell before merging.
        top_cell = ws.cell(start_row, col)
        if template_batch_top_row is not None:
            src_cell = ws.cell(template_batch_top_row, col)
            copy_cell_with_formula_translation(src_cell, top_cell)

        # Leave confirm column O blank, while keeping style.
        if col == CONFIRM_COL:
            top_cell.value = None

        # Clear lower cells before merge.
        for row in range(start_row + 1, end_row + 1):
            ws.cell(row, col).value = None

        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)


def apply_date_warning_if_needed(ws, row: int, is_continuous: bool) -> None:
    if not is_continuous:
        ws.cell(row=row, column=COL_START_DATE).fill = copy(DATE_WARNING_FILL)
        ws.cell(row=row, column=COL_END_DATE).fill = copy(DATE_WARNING_FILL)


def previous_batch_end_date(ws) -> Optional[date]:
    last_row = get_last_data_row(ws)
    if last_row < FIRST_DATA_ROW:
        return None
    return cell_date_value(ws.cell(row=last_row, column=COL_END_DATE))


def date_format_like_template(ws, template_row: int, target_row: int) -> None:
    ws.cell(target_row, COL_START_DATE).number_format = ws.cell(template_row, COL_START_DATE).number_format
    ws.cell(target_row, COL_END_DATE).number_format = ws.cell(template_row, COL_END_DATE).number_format


# =========================
# Main migration logic
# =========================

def extract_embedded_images(ws_pics) -> List[Image.Image]:
    images: List[Image.Image] = []
    for img in getattr(ws_pics, "_images", []):
        raw = img._data()
        pil_img = Image.open(io.BytesIO(raw)).convert("RGB")
        images.append(pil_img)
    return images


def group_results_by_period(results: Iterable[ImageResult]) -> List[Tuple[Tuple[date, date], List[ImageResult]]]:
    grouped: Dict[Tuple[date, date], List[ImageResult]] = {}
    for result in results:
        grouped.setdefault((result.start_date, result.end_date), []).append(result)

    sorted_items = sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1]))
    order_index = {account: i for i, account in enumerate(ACCOUNT_ORDER)}

    final_items = []
    for period, rows in sorted_items:
        rows.sort(key=lambda r: order_index.get(r.account, 999))
        final_items.append((period, rows))
    return final_items


def print_detected_results(results: List[ImageResult]) -> None:
    print("\nDetected image results:")
    print("=" * 80)
    for r in sorted(results, key=lambda x: (x.start_date, x.end_date, x.account)):
        amount_text = ", ".join(f"{amount:,}" for amount in r.amounts)
        date_text = ", ".join(d.strftime("%Y/%m/%d") for d in r.dates)
        print(f"Image {r.image_index}")
        print(f"  Account   : {r.account}")
        print(f"  Start date: {r.start_date.strftime('%Y/%m/%d')}")
        print(f"  End date  : {r.end_date.strftime('%Y/%m/%d')}")
        print(f"  Revenue   : {r.revenue:,}")
        print(f"  Dates     : {date_text}")
        print(f"  Amounts   : {amount_text}")
        print("-" * 80)


def update_or_append_reconciliation(ws, results: List[ImageResult]) -> None:
    existing = existing_row_map(ws)
    grouped = group_results_by_period(results)
    last_existing_end = previous_batch_end_date(ws)

    for (start, end), batch_results in grouped:
        expected_start = last_existing_end + timedelta(days=1) if last_existing_end else None
        is_continuous = expected_start is None or start == expected_start

        existing_rows_for_batch: List[int] = []
        missing_results: List[ImageResult] = []

        for result in batch_results:
            key = (result.account, result.start_date, result.end_date)
            row = existing.get(key)
            if row:
                ws.cell(row=row, column=COL_REVENUE).value = result.revenue
                apply_date_warning_if_needed(ws, row, is_continuous)
                existing_rows_for_batch.append(row)
            else:
                missing_results.append(result)

        appended_rows: List[int] = []
        if missing_results:
            template_row = get_last_data_row(ws)
            if template_row < FIRST_DATA_ROW:
                raise ValueError("No template data row found in the reconciliation sheet. Please keep at least one existing data row.")

            insert_start_row = template_row + 1
            template_batch_top_row = find_previous_batch_top_row(ws, insert_start_row)

            for i, result in enumerate(missing_results):
                target_row = insert_start_row + i
                copy_row_like_drag_down(ws, template_row, target_row)
                clear_row_transfer_area(ws, target_row)

                ws.cell(row=target_row, column=COL_ACCOUNT).value = result.account
                ws.cell(row=target_row, column=COL_START_DATE).value = result.start_date
                ws.cell(row=target_row, column=COL_END_DATE).value = result.end_date
                ws.cell(row=target_row, column=COL_REVENUE).value = result.revenue

                date_format_like_template(ws, template_row, target_row)
                apply_date_warning_if_needed(ws, target_row, is_continuous)

                existing[(result.account, result.start_date, result.end_date)] = target_row
                appended_rows.append(target_row)

            setup_transfer_area_for_batch(ws, min(appended_rows), max(appended_rows), template_batch_top_row)

        batch_rows = sorted(existing_rows_for_batch + appended_rows)
        if batch_rows:
            template_batch_top_row = find_previous_batch_top_row(ws, min(batch_rows))
            setup_transfer_area_for_batch(ws, min(batch_rows), max(batch_rows), template_batch_top_row)

        last_existing_end = end


def delete_images_from_sheet(ws_pics) -> None:
    ws_pics._images = []


def find_latest_workbook(folder: Path) -> Path:
    candidates = []
    for path in folder.glob(WORKBOOK_NAME_PATTERN):
        if path.name.startswith("~$"):
            continue
        if path.parent.name.lower() == LOG_FOLDER_NAME.lower():
            continue
        candidates.append(path)

    if not candidates:
        raise FileNotFoundError(
            f"No workbook found in {folder} matching pattern: {WORKBOOK_NAME_PATTERN}"
        )

    # Choose the newest modified workbook. This works with renamed timestamp workbooks.
    return max(candidates, key=lambda p: p.stat().st_mtime)


def make_timestamped_output_path(input_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    stem = input_path.stem

    # If the current workbook already ends with _yyyymmddhhmmss, remove that old suffix before adding a new one.
    stem = re.sub(r"_\d{14}$", "", stem)

    return input_path.with_name(f"{stem}_{timestamp}{input_path.suffix}")


def move_previous_workbook_to_log(input_path: Path, output_path: Path) -> None:
    log_folder = input_path.parent / LOG_FOLDER_NAME
    log_folder.mkdir(exist_ok=True)

    destination = log_folder / input_path.name
    if destination.exists():
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        destination = log_folder / f"{input_path.stem}_moved_{timestamp}{input_path.suffix}"

    # Do not move if input and output are somehow the same path.
    if input_path.resolve() != output_path.resolve():
        shutil.move(str(input_path), str(destination))
        print(f"Moved previous workbook to log folder:\n{destination}")


def migrate_workbook(input_path: Path) -> Path:
    output_path = make_timestamped_output_path(input_path)

    wb = load_workbook(input_path)
    print("Workbook sheets:", ", ".join(wb.sheetnames))

    ws_pics = get_sheet_by_alias(wb, SHEET_PICS_ALIASES)
    ws_recon = get_sheet_by_alias(wb, SHEET_RECONCILIATION_ALIASES)

    print(f"Using pics sheet          : {ws_pics.title}")
    print(f"Using reconciliation sheet: {ws_recon.title}")

    embedded_images = extract_embedded_images(ws_pics)
    if not embedded_images:
        raise ValueError(f"No embedded images found in sheet: {ws_pics.title}")

    results = [parse_image_result(idx, image) for idx, image in enumerate(embedded_images, start=1)]
    print_detected_results(results)

    update_or_append_reconciliation(ws_recon, results)
    delete_images_from_sheet(ws_pics)

    wb.save(output_path)
    print(f"\nDone. Saved updated workbook to:\n{output_path}")

    move_previous_workbook_to_log(input_path, output_path)
    return output_path


def main() -> None:
    current_folder = Path.cwd()

    if len(sys.argv) >= 2:
        input_path = Path(sys.argv[1]).expanduser().resolve()
    else:
        input_path = find_latest_workbook(current_folder).resolve()

    if not input_path.exists():
        raise FileNotFoundError(input_path)

    print(f"Selected workbook:\n{input_path}")
    output_path = migrate_workbook(input_path)
    print(f"\nMigration finished successfully. New working workbook:\n{output_path}")


if __name__ == "__main__":
    main()
