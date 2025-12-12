import os
import sys
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.worksheet.views import Selection

from config import (
    SCRAPE_FOLDER as CFG_SCRAPE_FOLDER,
    MAPPING_PATH as CFG_MAPPING_PATH,
)



# ======================== CONFIG ========================

# 从 config.py 读取统一配置
SCRAPE_FOLDER = CFG_SCRAPE_FOLDER
MAPPING_PATH = CFG_MAPPING_PATH

CODE_COL = "商品選項貨號"
URL_COL = "商品链接"
PID_COL = "商品ID"
ATTR_COL = "属性SKU"
SKU_ID_COL = "SKU ID"
SPEC_ID_COL = "Spec ID"
SHOP_COL = "店铺名称"

# 如果有警告或错误，则在退出前停一下
NEED_PAUSE = False


# ======================== Helpers ========================

def find_latest_done_workbook(folder: str) -> str:
    """在 folder 中查找最近修改的、文件名以 (done).xlsx 结尾的工作簿（忽略临时文件）。"""
    candidates = []
    for fname in os.listdir(folder):
        if fname.startswith("~$"):
            continue
        if not fname.lower().endswith(".xlsx"):
            continue
        name, ext = os.path.splitext(fname)
        if not name.endswith("(done)"):
            continue
        full = os.path.join(folder, fname)
        mtime = os.path.getmtime(full)
        candidates.append((mtime, full))

    if not candidates:
        raise FileNotFoundError(f"在 {folder} 下未找到任何以 (done).xlsx 结尾的文件。")

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


# ======================== Main Logic ========================

def fill_code_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    根据 商品ID + 样例行(商品選項貨號+属性SKU) 自动填充 商品選項貨號。
    无样例 或 属性SKU 为空 的行保持为空。
    """
    # 如果列名还是 Unnamed: 0，则重命名为 商品選項貨號
    if CODE_COL not in df.columns and "Unnamed: 0" in df.columns:
        df = df.rename(columns={"Unnamed: 0": CODE_COL})

    # 确保关键列存在
    for col in [CODE_COL, PID_COL, ATTR_COL]:
        if col not in df.columns:
            raise KeyError(f"工作簿缺少必要列: {col}")

    df[CODE_COL] = df[CODE_COL].astype(object)

    # 按 商品ID 分组
    for pid, group in df.groupby(PID_COL, dropna=False):
        idxs = group.index

        # 找样例行：商品選項貨號 非空 且 属性SKU 非空
        mask_sample = (
            group[CODE_COL].apply(lambda v: not is_empty(v)) &
            group[ATTR_COL].apply(lambda v: not is_empty(v))
        )

        if not mask_sample.any():
            # 没有样例，这一组全部保持原样
            continue

        sample_row = group[mask_sample].iloc[0]
        sample_name = str(sample_row[CODE_COL]).strip()
        sample_attr = str(sample_row[ATTR_COL]).strip()

        if sample_attr and sample_name.endswith(sample_attr):
            prefix = sample_name[:-len(sample_attr)]
        else:
            prefix = sample_name  # 兜底：不截取

        # 为该商品ID的其它行填充
        for idx in idxs:
            current_code = df.at[idx, CODE_COL]
            attr_val = df.at[idx, ATTR_COL]

            if not is_empty(current_code):
                # 已经有人为填写的，不覆盖
                continue
            if is_empty(attr_val):
                # 没有属性SKU，保持为空
                continue

            attr_str = str(attr_val).strip()
            df.at[idx, CODE_COL] = prefix + attr_str

    return df


def append_to_mapping(b_df: pd.DataFrame, mapping_path: str) -> pd.DataFrame:
    """
    从已经填好的 B(done).xlsx 中，筛选有用的行，追加到 Mapping_Data.xlsx。
    - 只追加 商品選項貨號 和 属性SKU 均非空的行
    - 不覆盖旧数据，只在末尾追加
    - 避免重复商品選項貨號
    """
    global NEED_PAUSE

    if not os.path.exists(mapping_path):
        raise FileNotFoundError(f"未找到 Mapping_Data.xlsx：{mapping_path}")

    map_df = pd.read_excel(mapping_path)

    if CODE_COL not in b_df.columns:
        raise KeyError(f"B(done) 缺少列: {CODE_COL}")
    required_b_cols = [URL_COL, PID_COL, ATTR_COL, SKU_ID_COL, SPEC_ID_COL, SHOP_COL]
    for col in required_b_cols:
        if col not in b_df.columns:
            raise KeyError(f"B(done) 缺少列: {col}")

    # 筛选：code 和 属性SKU 都非空的行
    def both_non_empty(row):
        return (not is_empty(row[CODE_COL])) and (not is_empty(row[ATTR_COL]))

    mask_use = b_df.apply(both_non_empty, axis=1)
    used_df = b_df[mask_use].copy()
    skipped_df = b_df[~mask_use].copy()

    print(f"[INFO] B(done) 总行数: {len(b_df)}")
    print(f"[INFO] 可用于 Mapping 的行数: {len(used_df)}")
    print(f"[WARN] 因 商品選項貨號 或 属性SKU 为空而跳过的行数: {len(skipped_df)}")
    if len(skipped_df) > 0:
        NEED_PAUSE = True  # 有警告，需要停留给你看
        print(skipped_df[[PID_COL, ATTR_COL, CODE_COL]].head(10))

    if used_df.empty:
        print("[INFO] 没有可追加到 Mapping_Data 的新数据。")
        return map_df

    # 避免重复：按 商品選項貨號 去重
    existing_codes = set(map_df[CODE_COL].astype(str))
    used_df[CODE_COL] = used_df[CODE_COL].astype(str)

    new_rows = used_df[~used_df[CODE_COL].isin(existing_codes)].copy()
    print(f"[INFO] 去掉已存在的商品選項貨號后，新追加行数: {len(new_rows)}")

    if new_rows.empty:
        print("[INFO] 所有行的 商品選項貨號 在 Mapping_Data 中已存在，不追加。")
        return map_df

    # 构造与 Mapping_Data 相同列结构的新 DataFrame
    new_map_df = pd.DataFrame({
        CODE_COL: new_rows[CODE_COL].astype(str),
        "商品链接": new_rows[URL_COL],
        "商品ID": new_rows[PID_COL],
        "属性SKU": new_rows[ATTR_COL],
        "SKU ID": new_rows[SKU_ID_COL],
        "Spec ID": new_rows[SPEC_ID_COL],
        "主供应商": new_rows[SHOP_COL],
        "属性SKU豁免": "",
        "商品链接.1": "",
        "商品ID.1": "",
        "属性SKU.1": "",
        "SKU ID.1": "",
        "Spec ID.1": "",
        "副供应商": "",
    })

    new_map_df = new_map_df.reindex(columns=map_df.columns, fill_value="")
    map_df = pd.concat([map_df, new_map_df], ignore_index=True)

    return map_df


def set_mapping_view_to_last_rows(path: str):
    """
    用 openpyxl 把 Mapping_Data.xlsx 的视图移动到最后几行，
    这样打开时光标就在新追加的行附近。
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        last_row = ws.max_row
        if last_row < 1:
            wb.save(path)
            return

        # 视图左上角定位到 A(last_row-20) 以便看到下面几行
        top_row = max(last_row - 20, 1)
        ws.sheet_view.topLeftCell = f"A{top_row}"

        # 选中最后一行的 A 列
        sel = list(ws.sheet_view.selection)
        target = f"A{last_row}"
        if sel:
            sel[0].activeCell = target
            sel[0].sqref = target
            ws.sheet_view.selection = tuple(sel)
        else:
            ws.sheet_view.selection = (Selection(activeCell=target, sqref=target),)

        wb.save(path)
    except Exception as e:
        # 这里不要让脚本失败，只是提醒
        print("[WARN] 设置 Mapping_Data 视图位置失败:", e)


def main():
    global NEED_PAUSE

    print("====================================================")
    print("  从 B(done).xlsx 填充 商品選項貨號 并追加到 Mapping_Data.xlsx")
    print("====================================================")
    print("工作目录:", SCRAPE_FOLDER)
    print("Mapping:", MAPPING_PATH)

    try:
        b_path = find_latest_done_workbook(SCRAPE_FOLDER)
    except FileNotFoundError as e:
        print(e)
        NEED_PAUSE = True
        return

    print(f"[INFO] 将处理最新的 (done) 工作簿: {b_path}")

    # 1) 读取 B(done).xlsx 并填充 商品選項貨號
    try:
        b_df = pd.read_excel(b_path)
    except Exception as e:
        print("[ERROR] 读取 B(done).xlsx 失败:", e)
        NEED_PAUSE = True
        return

    try:
        b_df = fill_code_column(b_df)
    except Exception as e:
        print("[ERROR] 填充 商品選項貨號 时出错:", e)
        NEED_PAUSE = True
        return

    try:
        b_df.to_excel(b_path, index=False)
        print(f"[INFO] 已写回更新后的 B(done) 文件: {b_path}")
    except Exception as e:
        print("[ERROR] 保存 B(done).xlsx 失败:", e)
        NEED_PAUSE = True
        return

    # 2) 追加到 Mapping_Data.xlsx
    try:
        map_df = append_to_mapping(b_df, MAPPING_PATH)
    except Exception as e:
        print("[ERROR] 处理 Mapping_Data 时出错:", e)
        NEED_PAUSE = True
        return

    # 保存 Mapping_Data，并把视图定位到最后几行
    try:
        map_df.to_excel(MAPPING_PATH, index=False)
        set_mapping_view_to_last_rows(MAPPING_PATH)
        print(f"[INFO] 已更新 Mapping_Data.xlsx: {MAPPING_PATH}")
    except Exception as e:
        print("[ERROR] 保存 Mapping_Data.xlsx 失败:", e)
        NEED_PAUSE = True
        return

    # 3) 打开两个文件
    try:
        os.startfile(b_path)
    except Exception as e:
        print("[WARN] 无法自动打开 B(done):", e)
        NEED_PAUSE = True

    try:
        os.startfile(MAPPING_PATH)
    except Exception as e:
        print("[WARN] 无法自动打开 Mapping_Data:", e)
        NEED_PAUSE = True

    print("全部处理完成。")

    # 如果过程中有警告或错误，则停一下给你看
    if NEED_PAUSE:
        input("存在警告或错误，请查看上方信息后按回车键退出...")


if __name__ == "__main__":
    main()
